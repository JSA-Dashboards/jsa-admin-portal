"""
import_history.py — Import historical basis data from Excel files into the basis tracker DB.

SUPPORTED LAYOUTS
-----------------
Place one or more .xlsx files in the  history/  folder and run this script.
The importer auto-detects two common layouts:

  Layout A — "Full bid sheet" (one row per delivery month)
  ┌────────────┬───────────────────────────────────┬──────────┬──────────────────┬──────────┬───────────────┐
  │ Date       │ Location                          │ Grain    │ Delivery         │ Futures  │ Basis (cents) │
  ├────────────┼───────────────────────────────────┼──────────┼──────────────────┼──────────┼───────────────┤
  │ 2025-01-15 │ ADM Decatur / Decatur IL Corn     │ Corn     │ Jan 2025         │ ZCH25    │ +10           │

  Layout B — "Spot-only" (just the cash basis for the day, no delivery month)
  ┌────────────┬───────────────────────────────────┬──────────┬───────────────┐
  │ Date       │ Location                          │ Grain    │ Basis (cents) │
  ├────────────┼───────────────────────────────────┼──────────┼───────────────┤
  │ 2025-01-15 │ ADM Decatur                       │ Corn     │ +10           │

Column names are case-insensitive and flexible — "date", "Date", "DATE" all work.
Basis values can be in cents ("10", "-35") or dollars ("0.10", "-0.35").
If the value looks like a dollar (abs < 5.0) it is auto-converted × 100.

LOCATION MAPPING
----------------
The importer normalises common abbreviations automatically, e.g.:
  "ADM CR" / "Cedar Rapids" / "ADM Cedar Rapids" → "Cedar Rapids, IA"
  "ADM Decatur Corn" / "Decatur Corn"            → "Decatur, IL (Corn Processing)"
  "ADM Decatur Soy" / "Decatur Soy"              → "Decatur, IL (Soy Processing)"
  "ADM St Louis" / "STL"                        → "St. Louis, MO (Elevator)"

Add custom mappings to the LOCATION_MAP dict below.

USAGE
-----
  python import_history.py --dry-run           # preview what would be imported
  python import_history.py                     # import to local SQLite
  python import_history.py --pg                # import to Supabase (DATABASE_URL required)
  python import_history.py --file my_data.xlsx # specific file
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import database as db

# ── Location normalisation map ─────────────────────────────────────────────────
# keys are lowercase; values are (provider, canonical_location)

LOCATION_MAP: dict[str, tuple[str, str]] = {
    # ADM Cedar Rapids variants
    "adm cedar rapids":            ("ADM", "Cedar Rapids, IA"),
    "cedar rapids":                ("ADM", "Cedar Rapids, IA"),
    "cedar rapids ia":             ("ADM", "Cedar Rapids, IA"),
    "adm cr":                      ("ADM", "Cedar Rapids, IA"),
    "cr":                          ("ADM", "Cedar Rapids, IA"),
    # ADM Decatur — Corn
    "adm decatur corn":            ("ADM", "Decatur, IL (Corn Processing)"),
    "decatur corn":                ("ADM", "Decatur, IL (Corn Processing)"),
    "adm decatur il corn":         ("ADM", "Decatur, IL (Corn Processing)"),
    "decatur corn processing":     ("ADM", "Decatur, IL (Corn Processing)"),
    "decatur il corn":             ("ADM", "Decatur, IL (Corn Processing)"),
    "decatur, il (corn processing)": ("ADM", "Decatur, IL (Corn Processing)"),
    # ADM Decatur — Soy
    "adm decatur soy":             ("ADM", "Decatur, IL (Soy Processing)"),
    "decatur soy":                 ("ADM", "Decatur, IL (Soy Processing)"),
    "adm decatur il soy":          ("ADM", "Decatur, IL (Soy Processing)"),
    "decatur soy processing":      ("ADM", "Decatur, IL (Soy Processing)"),
    "decatur il soy":              ("ADM", "Decatur, IL (Soy Processing)"),
    "decatur, il (soy processing)": ("ADM", "Decatur, IL (Soy Processing)"),
    # ADM Decatur — catch-all (old email name, ambiguous → corn+soy both)
    "adm decatur":                 ("ADM", "__DECATUR_BOTH__"),
    "decatur":                     ("ADM", "__DECATUR_BOTH__"),
    "decatur il":                  ("ADM", "__DECATUR_BOTH__"),
    "decatur, il":                 ("ADM", "__DECATUR_BOTH__"),
    # ADM St. Louis
    "adm st. louis":               ("ADM", "St. Louis, MO (Elevator)"),
    "adm st louis":                ("ADM", "St. Louis, MO (Elevator)"),
    "st. louis":                   ("ADM", "St. Louis, MO (Elevator)"),
    "st louis":                    ("ADM", "St. Louis, MO (Elevator)"),
    "stl":                         ("ADM", "St. Louis, MO (Elevator)"),
    "st. louis, mo (elevator)":    ("ADM", "St. Louis, MO (Elevator)"),
    "st. louis mo":                ("ADM", "St. Louis, MO (Elevator)"),
}

# Grain normalisation
GRAIN_MAP: dict[str, str] = {
    "corn": "Corn", "c": "Corn",
    "soybeans": "Soybeans", "soy": "Soybeans", "beans": "Soybeans", "s": "Soybeans",
    "wheat": "Wheat", "w": "Wheat",
    "soybean meal": "Soybean Meal", "meal": "Soybean Meal",
    "soybean oil": "Soybean Oil", "oil": "Soybean Oil",
}

# Month name → CME code mapping for symbol inference
_CME_MONTH: dict[str, str] = {
    "jan": "F", "feb": "G", "mar": "H", "apr": "J",
    "may": "K", "jun": "M", "jul": "N", "aug": "Q",
    "sep": "U", "oct": "V", "nov": "X", "dec": "Z",
}
_GRAIN_ROOT: dict[str, str] = {
    "Corn": "ZC", "Soybeans": "ZS", "Wheat": "ZW",
    "Soybean Meal": "ZM", "Soybean Oil": "ZL",
}


def _norm_loc(raw: str) -> tuple[str, str] | None:
    """Normalise a raw location string to (provider, canonical_location)."""
    key = raw.strip().lower()
    key = re.sub(r"\s+", " ", key)
    key = key.rstrip(".")
    if key in LOCATION_MAP:
        return LOCATION_MAP[key]
    # fuzzy: try stripping trailing state abbrev  "adm cedar rapids, ia" → "adm cedar rapids"
    key2 = re.sub(r",?\s*[a-z]{2}$", "", key)
    if key2 in LOCATION_MAP:
        return LOCATION_MAP[key2]
    return None


def _norm_grain(raw: str) -> str | None:
    key = raw.strip().lower()
    return GRAIN_MAP.get(key)


def _infer_cme(grain: str, delivery: str) -> str | None:
    """
    Infer a CME futures symbol from grain + delivery string like "Jan 2025".
    Returns None if it can't be determined.
    """
    root = _GRAIN_ROOT.get(grain)
    if not root:
        return None
    m = re.search(r"([a-z]{3})\w*\s*'?(\d{2,4})", delivery.lower())
    if not m:
        return None
    mon_code = _CME_MONTH.get(m.group(1)[:3])
    if not mon_code:
        return None
    yr = int(m.group(2))
    if yr > 100:
        yr = yr % 100
    return f"{root}{mon_code}{yr:02d}"


def _parse_basis(raw) -> int | None:
    """Parse a basis value to integer cents. Handles cents or dollar notation."""
    if raw is None:
        return None
    txt = str(raw).strip().replace(",", "")
    if txt in ("", "N/A", "TBD", "-", "—"):
        return None
    m = re.search(r"([+-]?\d+\.?\d*)", txt)
    if not m:
        return None
    try:
        val = float(m.group(1))
        # Heuristic: if abs < 5.0, treat as dollars → convert to cents
        if abs(val) < 5.0:
            return round(val * 100)
        return int(val)
    except ValueError:
        return None


def _parse_date(raw) -> str | None:
    """Return ISO date string like '2025-01-15T00:00:00Z', or None."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%dT00:00:00Z")
    # Try common string formats
    txt = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%b-%y", "%B %d, %Y"):
        try:
            return datetime.strptime(txt, fmt).strftime("%Y-%m-%dT00:00:00Z")
        except ValueError:
            continue
    return None


# ── Column name detection ─────────────────────────────────────────────────────

_DATE_COLS     = {"date", "day", "trade date", "price date", "report date"}
_LOC_COLS      = {"location", "loc", "facility", "plant", "elevator", "provider"}
_GRAIN_COLS    = {"grain", "commodity", "crop"}
_DELIVERY_COLS = {"delivery", "delivery month", "period", "month", "contract",
                  "delivery period", "future period"}
_FUTURES_COLS  = {"futures", "futures symbol", "symbol", "cme symbol", "contract symbol"}
_BASIS_COLS    = {"basis", "basis (cents)", "basis cents", "basis¢", "cash basis",
                  "net basis", "flat price basis", "spot basis"}


def _find_col(headers: list[str], candidates: set[str]) -> str | None:
    for h in headers:
        if h.strip().lower() in candidates:
            return h
    return None


# ── Excel reading ──────────────────────────────────────────────────────────────

def _read_excel(path: Path) -> list[dict]:
    """Return list of row dicts from all sheets in the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Find header row (first non-empty row)
        header_idx = 0
        for i, row in enumerate(rows):
            if any(c is not None for c in row):
                header_idx = i
                break
        headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        for row in rows[header_idx + 1:]:
            if all(c is None for c in row):
                continue
            all_rows.append({"_sheet": sheet_name,
                             **{headers[i]: row[i] for i in range(len(headers))}})
    wb.close()
    return all_rows


# ── Row parsing ────────────────────────────────────────────────────────────────

def _parse_rows(raw_rows: list[dict], source_name: str) -> list[dict]:
    """
    Parse raw Excel rows into a list of structured dicts ready for the DB.
    Returns list of:
    {
        "timestamp":     str,
        "provider":      str,
        "location":      str,
        "grain":         str,
        "delivery":      str,       # "" for spot-only
        "futures":       str,       # "" if not available
        "basis_cents":   int,
        "is_spot":       bool,
        "_source_row":   int,
    }
    """
    if not raw_rows:
        return []

    headers = list(raw_rows[0].keys())
    headers = [h for h in headers if h != "_sheet"]

    date_col     = _find_col(headers, _DATE_COLS)
    loc_col      = _find_col(headers, _LOC_COLS)
    grain_col    = _find_col(headers, _GRAIN_COLS)
    delivery_col = _find_col(headers, _DELIVERY_COLS)
    futures_col  = _find_col(headers, _FUTURES_COLS)
    basis_col    = _find_col(headers, _BASIS_COLS)

    # If no basis column matched, try last numeric-looking column
    if basis_col is None:
        for h in reversed(headers):
            if h.lower() not in (_DATE_COLS | _LOC_COLS | _GRAIN_COLS | _DELIVERY_COLS | _FUTURES_COLS):
                basis_col = h
                break

    if not date_col or not basis_col:
        print(f"  WARNING [{source_name}]: couldn't find date or basis column in {headers}")
        return []

    parsed = []
    for i, row in enumerate(raw_rows, start=2):
        date_ts = _parse_date(row.get(date_col))
        if not date_ts:
            continue

        basis = _parse_basis(row.get(basis_col))
        if basis is None:
            continue

        # Location
        raw_loc = str(row.get(loc_col, "") or "").strip()
        loc_mapped = _norm_loc(raw_loc)
        if not loc_mapped:
            print(f"  WARNING row {i}: unrecognised location {raw_loc!r} — skipping")
            continue
        provider, location = loc_mapped

        # Grain
        raw_grain = str(row.get(grain_col, "") or "").strip()
        grain = _norm_grain(raw_grain) if raw_grain else None
        if not grain:
            # Infer from location name
            if "corn" in location.lower():
                grain = "Corn"
            elif "soy" in location.lower():
                grain = "Soybeans"
            else:
                print(f"  WARNING row {i}: can't determine grain from {raw_grain!r} — skipping")
                continue

        # Delivery
        raw_delivery = str(row.get(delivery_col, "") or "").strip() if delivery_col else ""
        is_spot = not raw_delivery or raw_delivery.lower() in ("spot", "cash", "")

        # Futures symbol
        raw_futures = str(row.get(futures_col, "") or "").strip() if futures_col else ""
        if not raw_futures and not is_spot:
            raw_futures = _infer_cme(grain, raw_delivery) or ""

        # Handle __DECATUR_BOTH__: fan out by grain
        if location == "__DECATUR_BOTH__":
            if grain == "Corn":
                location = "Decatur, IL (Corn Processing)"
            elif grain == "Soybeans":
                location = "Decatur, IL (Soy Processing)"
            else:
                location = "Decatur, IL (Corn Processing)"  # fallback

        parsed.append({
            "timestamp":   date_ts,
            "provider":    provider,
            "location":    location,
            "grain":       grain,
            "delivery":    raw_delivery,
            "futures":     raw_futures,
            "basis_cents": basis,
            "is_spot":     is_spot,
            "_source_row": i,
        })

    return parsed


# ── DB import ─────────────────────────────────────────────────────────────────

def _group_into_snapshots(rows: list[dict]) -> list[dict]:
    """
    Group parsed rows into snapshot dicts keyed by (timestamp, provider, location).
    Returns list of snapshot dicts ready for upsert_snapshot().
    """
    from collections import defaultdict
    groups: dict[tuple, list] = defaultdict(list)
    for r in rows:
        key = (r["timestamp"], r["provider"], r["location"])
        groups[key].append(r)

    snaps = []
    for (ts, prov, loc), loc_rows in groups.items():
        snap_rows = []
        seen: set[str] = set()
        for r in loc_rows:
            if r["is_spot"]:
                row_id = f"HIST_{r['grain'][:3].upper()}_SPOT"
            else:
                delivery_clean = re.sub(r"[^A-Za-z0-9]", "", r["delivery"])
                row_id = f"HIST_{r['grain'][:3].upper()}_{delivery_clean}"

            # Deduplicate within snapshot
            suffix = 0
            base_id = row_id
            while row_id in seen:
                suffix += 1
                row_id = f"{base_id}_{suffix}"
            seen.add(row_id)

            snap_rows.append({
                "id":            row_id,
                "grain":         r["grain"],
                "deliveryMonth": r["delivery"] if not r["is_spot"] else r["grain"],
                "futuresSymbol": r["futures"],
                "basisCents":    r["basis_cents"],
                "isSpot":        r["is_spot"],
                "spotGrain":     r["grain"] if r["is_spot"] else None,
            })

        snaps.append({
            "timestamp": ts,
            "provider":  prov,
            "location":  loc,
            "source":    "historical",
            "rows":      snap_rows,
        })
    return snaps


def import_file(path: Path, dry_run: bool = True, verbose: bool = True) -> dict:
    """Import one Excel file. Returns {"snapshots": int, "rows": int, "skipped": int}."""
    print(f"\n  File: {path.name}")
    raw_rows = _read_excel(path)
    if not raw_rows:
        print("  (empty)")
        return {"snapshots": 0, "rows": 0, "skipped": 0}

    parsed = _parse_rows(raw_rows, path.name)
    if not parsed:
        print("  (no valid rows parsed)")
        return {"snapshots": 0, "rows": 0, "skipped": 0}

    snaps = _group_into_snapshots(parsed)

    if dry_run:
        # Show summary by location
        from collections import defaultdict
        loc_counts: dict[str, int] = defaultdict(int)
        for s in snaps:
            loc_counts[f"{s['provider']} / {s['location']}"] += 1
        print(f"  Parsed {len(parsed)} rows → {len(snaps)} snapshot(s)")
        for loc_key, n in sorted(loc_counts.items()):
            print(f"    {n:4d} snapshot(s)  {loc_key}")
        # Show a few sample rows
        if verbose and parsed:
            print("  Sample rows:")
            for r in parsed[:5]:
                spot_str = "SPOT" if r["is_spot"] else r["delivery"]
                print(f"    {r['timestamp'][:10]}  {r['provider']}/{r['location']:35s}  "
                      f"{r['grain']:10s}  {spot_str:20s}  {r['basis_cents']:+d}c")
        return {"snapshots": len(snaps), "rows": len(parsed), "skipped": 0}

    # Apply
    imported_snaps = 0
    imported_rows  = 0
    for snap in snaps:
        db.upsert_snapshot(snap)
        imported_snaps += 1
        imported_rows  += len(snap["rows"])

    print(f"  Imported {imported_snaps} snapshot(s), {imported_rows} row(s)")
    return {"snapshots": imported_snaps, "rows": imported_rows, "skipped": 0}


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import historical basis data from Excel files")
    parser.add_argument("--file", type=str, default=None,
                        help="Path to a specific .xlsx file (default: all files in history/)")
    parser.add_argument("--dry-run", action="store_true", default=False,
                        help="Preview what would be imported without writing to DB (default)")
    parser.add_argument("--apply", action="store_true", default=False,
                        help="Actually write to the database")
    parser.add_argument("--pg", action="store_true",
                        help="Use DATABASE_URL / Supabase instead of local SQLite")
    parser.add_argument("--quiet", action="store_true",
                        help="Reduce output verbosity")
    args = parser.parse_args()

    if args.pg and not os.getenv("DATABASE_URL"):
        print("ERROR: --pg requires DATABASE_URL to be set.")
        sys.exit(1)

    dry_run = not args.apply  # default is dry run
    verbose = not args.quiet

    backend = "PostgreSQL" if db._use_pg() else "SQLite"
    print(f"\nBasis Tracker — Historical Import  [{backend}]")
    print(f"Mode: {'DRY RUN' if dry_run else 'APPLY'}")
    print("=" * 60)

    db.init_db()

    if args.file:
        files = [Path(args.file)]
    else:
        history_dir = Path(__file__).parent / "history"
        if not history_dir.exists():
            print(f"\nNo 'history/' folder found at {history_dir}")
            print("Create it and place your .xlsx files inside, then re-run.")
            sys.exit(0)
        files = sorted(history_dir.glob("*.xlsx"))
        if not files:
            print(f"\nNo .xlsx files found in {history_dir}")
            sys.exit(0)

    total_snaps = 0
    total_rows  = 0
    for f in files:
        result = import_file(f, dry_run=dry_run, verbose=verbose)
        total_snaps += result["snapshots"]
        total_rows  += result["rows"]

    print()
    print("=" * 60)
    if dry_run:
        print(f"DRY RUN: would import {total_snaps} snapshot(s), {total_rows} row(s)")
        print("Re-run with --apply to commit.")
    else:
        print(f"Done: imported {total_snaps} snapshot(s), {total_rows} row(s)")


if __name__ == "__main__":
    main()
