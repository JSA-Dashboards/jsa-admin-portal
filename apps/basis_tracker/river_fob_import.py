"""
On-demand River FOB import — parse a JSA FOB Sheet workbook and return dated
snapshots for the "Update from the FOB sheet" control on the River FOB tab.

Self-contained copy of the River FOB portal's adaptive parser (import_history.
parse_tab): auto-detects the label column, reads each tab's own month/contract
headers so a month-turn roll is captured, normalizes freight regions. Works on a
local workbook path OR an uploaded file (bytes), so it functions on the deployed
cloud app too. Writes go through river_fob_data.save_snapshot (shared Supabase).
"""
import os
import re
import glob
import datetime as dt

import openpyxl

FOB_ROOT = os.environ.get("FOB_ROOT") or (
    r"C:\Users\KoltenPostin\John Stewart and Associates"
    r"\JSA - Documents\St. Louis\JSA FOB Sheet")

MONTH_NAME = ("January February March April May June July August September "
              "October November December").split()
MONTHS_RE = re.compile(r"JSA FOB Sheet.*(" + "|".join(MONTH_NAME) + r")\s*(\d{4})", re.I)
DATED_TAB = re.compile(r"\d{1,2}[-.]\d{1,2}([-.]\d{2,4})?")

COMMODITY_ALIASES = {"corn": "Corn", "soybeans": "Soybeans", "soybean": "Soybeans",
                     "beans": "Soybeans", "wheat": "Wheat"}
REGION_ALIASES = {
    "lower miss": "Lower Miss", "davenport south": "Davenport South",
    "mcgregor south": "McGregor South", "upper miss": "Upper Miss",
    "ohio": "Ohio", "stl": "STL", "il": "IL",
}
_MNUM = {"jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
         "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
         "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
         "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12}
_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul",
         8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
# The live archive + basis-tracker lookups use the sheet's June/July full form.
_MFIX = {"Jun": "June", "Jul": "July"}


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _tab_date(tab, wb_year, wb_month):
    parts = re.split(r"[-.]", tab.strip())
    try:
        mo, day = int(parts[0]), int(parts[1])
    except (ValueError, IndexError):
        return None
    if len(parts) >= 3 and parts[2]:
        yy = int(parts[2]); year = 2000 + yy if yy < 100 else yy
    else:
        year = wb_year
        if wb_month == 12 and mo == 1:
            year += 1
        elif wb_month == 1 and mo == 12:
            year -= 1
    try:
        return dt.date(year, mo, day)
    except ValueError:
        return None


def _norm_region(label):
    s = label[:-len("Freight")].strip()
    if s.lower().endswith("barge"):
        s = s[:-len("barge")].strip()
    return REGION_ALIASES.get(s.lower())


def _canon_month(label):
    s = str(label).strip().lower().rstrip(".")
    if s == "spot":
        return "Spot"
    if s[:2] in ("fh", "lh") or s in ("tw", "nw", ""):
        return None
    return _ABBR.get(_MNUM.get(s))


def parse_tab(ws):
    """Return (cif, freight, calendar) for one dated worksheet, or None."""
    a = sum(1 for r in range(1, 60) if "FOB Barge" in str(ws.cell(r, 1).value or ""))
    b = sum(1 for r in range(1, 60) if "FOB Barge" in str(ws.cell(r, 2).value or ""))
    lc = 1 if a >= b else 2
    if max(a, b) == 0:
        return None

    rows = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, lc).value
        if isinstance(v, str) and v.strip():
            rows.setdefault(r, v.strip())

    sections = []
    for r, lbl in rows.items():
        key = COMMODITY_ALIASES.get(lbl.strip().lower())
        if key:
            sections.append((r, key))
    sections.sort()
    if not sections:
        return None
    bounds = [(sections[i][0], sections[i][1],
               sections[i + 1][0] if i + 1 < len(sections) else ws.max_row + 1)
              for i in range(len(sections))]

    cif, freight, calendar = {}, {}, {}
    for start, commodity, end in bounds:
        cbot_r = next((r for r in range(start, end)
                       if str(ws.cell(r, lc).value or "").strip().upper() == "CBOT"), None)
        if not cbot_r:
            continue
        header_r, contract_r = cbot_r - 2, cbot_r - 1

        def _collect(candidate_cols, contiguous=False):
            dcols, mons, cons, seen = [], [], [], set()
            started = False
            for c in candidate_cols:
                lbl = _canon_month(ws.cell(header_r, c).value)
                if lbl is None or lbl in seen:
                    if contiguous and started:
                        break
                    continue
                started = True
                seen.add(lbl)
                dcols.append(c)
                mons.append(lbl)
                cons.append(str(ws.cell(contract_r, c).value or "").strip())
            return dcols, mons, cons

        cbot_cols = [c for c in range(lc + 1, ws.max_column + 1)
                     if isinstance(ws.cell(cbot_r, c).value, (int, float))]
        data_cols, months, contracts = _collect(cbot_cols)
        h_cols, h_months, h_contracts = _collect(
            range(lc + 1, ws.max_column + 1), contiguous=True)
        if len(h_cols) > len(data_cols):
            data_cols, months, contracts = h_cols, h_months, h_contracts
        if not data_cols:
            continue
        calendar[commodity] = list(zip(months, contracts))

        for r in range(start, end):
            lbl = str(ws.cell(r, lc).value or "").strip()
            if lbl.upper() == "CIF":
                cif[commodity] = {months[i]: _num(ws.cell(r, c).value)
                                  for i, c in enumerate(data_cols)}
            elif lbl.endswith("Freight"):
                reg = _norm_region(lbl)
                if reg and reg not in freight:
                    freight[reg] = {months[i]: _num(ws.cell(r, c).value)
                                    for i, c in enumerate(data_cols)}
    if not cif:
        return None
    return cif, freight, calendar


def find_active_workbook(today=None):
    """Newest 'JSA FOB Sheet … <Month> <Year>.xlsx' for the current month in the
    year folder (fallback: previous month). None if the folder isn't reachable."""
    today = today or dt.date.today()
    folder = os.path.join(FOB_ROOT, str(today.year))
    if not os.path.isdir(folder):
        return None

    def _cands(month_name, yr):
        out = []
        for f in glob.glob(os.path.join(folder, "*.xlsx")):
            b = os.path.basename(f)
            if b.startswith("~$"):
                continue
            m = MONTHS_RE.search(b)
            if m and m.group(1).lower() == month_name.lower() and int(m.group(2)) == yr:
                out.append(f)
        return sorted(out, key=os.path.getmtime, reverse=True)

    c = _cands(today.strftime("%B"), today.year)
    if not c:
        prev = today.replace(day=1) - dt.timedelta(days=1)
        c = _cands(prev.strftime("%B"), prev.year)
    return c[0] if c else None


def _fix(mv):
    return {_MFIX.get(k, k): v for k, v in mv.items()}


def _fixcal(cols):
    return [(_MFIX.get(m, m), c) for m, c in cols]


def import_workbook(src, name="", recent=8):
    """Parse the most recent `recent` dated tabs of a workbook.

    src  : a filesystem path OR a file-like/bytes object (uploaded file).
    name : the workbook filename (for month/year inference on 'M.D' tabs).
    Returns [(as_of_iso, cif, freight, calendar), …], newest last, with the
    Jun/Jul month labels remapped to the archive's June/July form.
    """
    m = MONTHS_RE.search(name or (src if isinstance(src, str) else ""))
    if m:
        wb_month = MONTH_NAME.index(m.group(1).capitalize()) + 1
        wb_year = int(m.group(2))
    else:
        t = dt.date.today()
        wb_month, wb_year = t.month, t.year

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    dated = []
    for tab in wb.sheetnames:
        if not DATED_TAB.fullmatch(tab.strip()):
            continue
        d = _tab_date(tab, wb_year, wb_month)
        if d:
            dated.append((d, tab))
    dated.sort()

    out = []
    for d, tab in dated[-recent:]:
        res = parse_tab(wb[tab])
        if not res:
            continue
        cif, freight, calendar = res
        cif = {c: _fix(mv) for c, mv in cif.items()}
        freight = {r: _fix(mv) for r, mv in freight.items()}
        calendar = {c: _fixcal(cols) for c, cols in calendar.items()}
        out.append((d.isoformat(), cif, freight, calendar))
    return out
