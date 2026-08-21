"""
import_soybean_history.py — Import weekly soybean spot basis history from
"Corn basis history .xlsx" (Soybeans sheet)

Sheet: Soybeans
  Rows 0-3 : Headers (Region / Company / Location / State)
  Row 4+   : Data
  Col 0    : Date
  Col 1    : Opt Mo  (CME month code: F/H/K/N/Q/U/X)
  Col 2    : Futures price (ignored)
  Col 3    : Reference basis (ignored)
  Col 4+   : One location per column, basis in cents, "Closed"/blank = skip

Only the columns that match currently-scraped locations are imported.
All other columns are ignored.  Grain = Soybeans, spot basis only.

Usage:
    python import_soybean_history.py               # dry run — shows counts only
    python import_soybean_history.py --apply       # write to local SQLite
    python import_soybean_history.py --apply --pg  # write to Supabase
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import database as db

# ── Header-based mapping ─────────────────────────────────────────────────────
# Maps the spreadsheet's own header (Company, Location, State) → the canonical
# (provider, db_location) used by the live scrapers, so historical and live data
# line up.  Keyed by header text (not column index) so inserting/moving columns
# in the spreadsheet never silently mis-maps the import.
# Spreadsheet header rows: Row0=Region, Row1=Company, Row2=Location, Row3=State.

NAME_MAP: dict[tuple[str, str, str], tuple[str, str]] = {
    # ── West region ──────────────────────────────────────────────────────────
    ("AGP", "Eagle Grove", "IA"):     ("AGP",        "AGP Eagle Grove, IA"),
    ("AGP", "Mason City", "IA"):      ("AGP",        "AGP Mason City, IA"),
    ("Cargill", "Cedar Rapids", "IA"):("Cargill",    "Cedar Rapids East"),
    ("Cargill", "Iowa Falls", "IA"):  ("Cargill",    "Iowa Falls"),
    ("Shell Rock", "Shell Rock", "IA"):("ShellRock", "Shell Rock"),
    ("ADM", "Des Moines", "IA"):      ("ADM",        "Des Moines, IA"),
    ("AGP", "Sheldon", "IA"):         ("AGP",        "AGP Sheldon, IA"),
    ("AGP", "E'burg", "IA"):          ("AGP",        "AGP Emmetsburg, IA"),  # E'burg = Emmetsburg
    ("CHS", "Fairmont", "MN"):        ("CHS",        "Fairmont"),
    ("AGP", "Dawson", "MN"):          ("AGP",        "AGP Dawson, MN"),
    ("CHS", "Mankato", "MN"):         ("CHS",        "Mankato"),
    ("ADM", "Mankato", "MN"):         ("ADM",        "Mankato, MN (Soy Processing)"),
    ("Cargill", "Sioux City", "IA"):  ("Cargill",    "Sioux City"),
    ("AGP", "SGT Bluff", "IA"):       ("AGP",        "AGP Sergeant Bluff, IA"),
    ("AGP", "Manning", "IA"):         ("AGP",        "AGP Manning, IA"),
    ("Platinum", "Alta", "IA"):       ("Platinum",   "Alta"),
    ("AGP", "St Joe", "MO"):          ("AGP",        "AGP St. Joseph, MO"),
    ("ADM", "Lincoln", "NE"):         ("ADM",        "Lincoln, NE (Soy Processing)"),
    ("ADM", "Fremont", "NE"):         ("ADM",        "Fremont, NE (Soy Processing)"),
    ("Norfolk", "Norfolk", "NE"):     ("NorfolkCrush","Norfolk"),
    ("MNSP", "Brewster", "MN"):       ("MNSP",       "Brewster"),
    ("Bunge", "Council Bluffs", "IA"):("Bunge",      "Council Bluffs, IA"),
    ("AGP", "Hastings", "NE"):        ("AGP",        "AGP Hastings, NE - Soy Plant"),
    ("AGP", "David City", "SD"):      ("AGP",        "AGP David City, NE"),
    ("AGP", "Aberdeen", "SD"):        ("AGP",        "AGP Aberdeen, SD"),
    ("SDSP", "Volga", "SD"):          ("SDSP",       "Volga"),     # South Dakota Soybean Processors
    ("High Plains", "Mitchell", "SD"):("HPPSD",      "Mitchell"),  # High Plains Processing, Mitchell SD
    ("CGB", "Casselton", "ND"):       ("NDSP",       "Casselton"), # North Dakota Soybean Processors
    ("ADM", "Spiritwood", "ND"):      ("ADM",        "Green Bison Soy Processing"),
    ("Bunge", "Emporia", "KS"):       ("Bunge",      "Emporia, KS"),
    ("Cargill", "KC", "MO"):          ("Cargill",    "Kansas City"),
    ("Cargill", "Wichita", "KS"):     ("Cargill",    "Wichita"),
    ("ADM", "Deerfield", "MO"):       ("ADM",        "Deerfield, MO"),
    ("ADM", "Mexico", "MO"):          ("ADM",        "Mexico, MO"),
    ("Bartlett", "Cherryvale", "KS"): ("Bartlett",   "Cherryvale"),

    # ── East region ──────────────────────────────────────────────────────────
    ("ADM", "Quincy", "IL"):          ("ADM",        "Quincy, IL (Soy Processing)"),
    ("Cargill", "Bloomington", "IL"): ("Cargill",    "Bloomington"),
    ("Bunge", "Gibson City", "IL"):   ("Bunge",      "Gibson City"),  # Bunge processor (no live scrape yet)
    ("Cargill", "Owensboro", "KY"):   ("Cargill",    "Owensboro"),
    ("ADM", "Fostoria", "OH"):        ("ADM",        "Fostoria, OH"),
    ("ADM", "Frankfort", "IN"):       ("ADM",        "Frankfort, IN"),
    ("ADM", "Decatur", "IL"):         ("ADM",        "Decatur, IL (Soy Processing)"),
    ("Cargill", "Lafayette", "IN"):   ("Cargill",    "Lafayette"),
    ("Cargill", "Sidney", "OH"):      ("Cargill",    "Sidney"),
    ("Bunge", "Morristown", "IN"):    ("Bunge",      "Morristown, IN - Bean Plant"),
    ("Bunge", "Decatur", "IN"):       ("Bunge",      "Decatur, IN"),
    ("Bunge", "Delphos", "OH"):       ("Bunge",      "Delphos, OH"),
    ("Bunge", "Bellevue", "OH"):      ("Bunge",      "Bellevue, OH"),
    ("CGB", "Mt Vernon", "IN"):       ("CGB",        "CGB MT VERNON"),
    ("LDC", "Claypool", "IN"):        ("LDC",        "Claypool"),
    ("White River", "Seymour", "IN"): ("WhiteRiver", "Seymour"),
    ("INCO", "Gilman", "IL"):         ("Cargill",    "Gilman"),   # formerly INCO; now Cargill
    ("Zeeland", "Zeeland", "MI"):     ("ZFS",        "Zeeland"),
    ("ZFS", "Ithaca", "MI"):          ("ZFS",        "ZFS Ithaca"),
    ("Cargill", "Guntersville", "AL"):("Cargill",    "Guntersville"),
    ("Bunge", "Decatur", "AL"):       ("Bunge",      "Decatur, AL"),
}

# Normalised (case/space-insensitive) lookup so minor header-typo drift still matches.
def _norm(s: str) -> str:
    return " ".join(str(s).strip().lower().split())

_NAME_MAP_NORM = {
    (_norm(a), _norm(b), _norm(c)): v for (a, b, c), v in NAME_MAP.items()
}


def build_col_map(rows: list) -> tuple[dict[int, tuple[str, str]], list]:
    """
    Map current spreadsheet columns → (provider, db_location) by matching the
    Company/Location/State header (rows 1/2/3) against NAME_MAP.  Returns
    (col_map, unmapped) where unmapped lists (col, company, location, state) for
    any data column we don't recognise (those are skipped, with a warning).
    """
    def hdr(ri: int, ci: int) -> str:
        return (str(rows[ri][ci]).strip()
                if ci < len(rows[ri]) and rows[ri][ci] not in (None, "") else "")

    ncol = max(len(r) for r in rows[:4])
    col_map: dict[int, tuple[str, str]] = {}
    unmapped: list = []
    for ci in range(4, ncol):
        comp, loc, st = hdr(1, ci), hdr(2, ci), hdr(3, ci)
        if not (comp or loc):
            continue
        canon = _NAME_MAP_NORM.get((_norm(comp), _norm(loc), _norm(st)))
        if canon:
            col_map[ci] = canon
        else:
            unmapped.append((ci, comp, loc, st))
    return col_map, unmapped

# CME month code → calendar month number
_CME_MONTH_NUM: dict[str, int] = {
    "F": 1, "G": 2, "H": 3, "J": 4, "K": 5, "M": 6,
    "N": 7, "Q": 8, "U": 9, "V": 10, "X": 11, "Z": 12,
}


def _infer_cme_soy(opt_mo: str, date: datetime) -> str:
    """
    Build a full CME soybean symbol from the Opt Mo letter and the row date.
    e.g.  "X" + 2024-09-02  →  "ZSX24"
          "F" + 2024-12-16  →  "ZSF25"  (January hasn't arrived yet in Dec)
    """
    code = opt_mo.strip().upper()
    month_num = _CME_MONTH_NUM.get(code)
    if not month_num:
        return ""
    year = date.year
    if month_num < date.month:
        year += 1
    return f"ZS{code}{year % 100:02d}"


def _parse_basis(raw) -> int | None:
    """Return integer cents or None.  Skips 'Closed', blank, non-numeric."""
    if raw is None:
        return None
    txt = str(raw).strip()
    if not txt or txt.lower() in ("closed", "n/a", "-", "—", "tbd"):
        return None
    try:
        return int(round(float(txt)))
    except (ValueError, TypeError):
        return None


def _upsert_one(conn, c, is_pg: bool, timestamp: str, provider: str, location: str,
                futures_sym: str, basis: int) -> None:
    """Insert a single spot snapshot using an already-open connection."""
    if is_pg:
        c.execute(
            """INSERT INTO snapshots (timestamp, provider, location, source)
               VALUES (%s, %s, %s, %s)
               ON CONFLICT (timestamp, provider, location) DO NOTHING
               RETURNING id""",
            (timestamp, provider, location, "historical"),
        )
        row = c.fetchone()
        if row:
            snap_id = row["id"]
        else:
            c.execute(
                "SELECT id FROM snapshots WHERE timestamp=%s AND provider=%s AND location=%s",
                (timestamp, provider, location),
            )
            snap_id = c.fetchone()["id"]
        c.execute(
            """INSERT INTO snapshot_rows
               (snapshot_id, row_id, grain, delivery_month, futures_symbol,
                basis_cents, is_spot, spot_grain)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
               ON CONFLICT (snapshot_id, row_id) DO NOTHING""",
            (snap_id, "HIST_SOY_SPOT", "Soybeans", "Soybeans", futures_sym, basis, 1, "Soybeans"),
        )
    else:
        c.execute(
            """INSERT OR IGNORE INTO snapshots (timestamp, provider, location, source)
               VALUES (?, ?, ?, ?)""",
            (timestamp, provider, location, "historical"),
        )
        if c.lastrowid == 0:
            c.execute(
                "SELECT id FROM snapshots WHERE timestamp=? AND provider=? AND location=?",
                (timestamp, provider, location),
            )
            snap_id = c.fetchone()["id"]
        else:
            snap_id = c.lastrowid
        c.execute(
            """INSERT OR IGNORE INTO snapshot_rows
               (snapshot_id, row_id, grain, delivery_month, futures_symbol,
                basis_cents, is_spot, spot_grain)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (snap_id, "HIST_SOY_SPOT", "Soybeans", "Soybeans", futures_sym, basis, 1, "Soybeans"),
        )


def run(file_path: Path, apply: bool = False, pg: bool = False) -> None:
    import openpyxl

    backend = "PostgreSQL" if db._use_pg() else "SQLite"
    print(f"\nSoybean History Import  [{backend}]  mode={'APPLY' if apply else 'DRY RUN'}")
    print("=" * 60)
    print(f"File: {file_path.name}")

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    if "Soybeans" not in wb.sheetnames:
        print(f"ERROR: sheet 'Soybeans' not found. Available: {wb.sheetnames}")
        wb.close()
        return
    ws = wb["Soybeans"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Rows 0-3 are headers; data starts at row index 4
    data_rows = rows[4:]

    # Build column → (provider, location) map from the header text (shift-proof).
    col_map, unmapped = build_col_map(rows)
    print(f"Mapped {len(col_map)} location column(s) from headers.")
    if unmapped:
        print("WARNING: unrecognised columns (skipped — add to NAME_MAP if needed):")
        for ci, comp, loc, st in unmapped:
            print(f"   col {ci}: {comp} | {loc} | {st}")

    counts: dict[tuple, int] = {v: 0 for v in col_map.values()}
    skipped = 0
    total_snaps = 0

    db.init_db()

    pending: list[tuple] = []

    for row in data_rows:
        raw_date = row[0]
        if raw_date is None:
            continue
        if isinstance(raw_date, datetime):
            date = raw_date
        else:
            try:
                date = datetime.strptime(str(raw_date).strip()[:10], "%Y-%m-%d")
            except ValueError:
                skipped += 1
                continue

        # Skip stray future-dated rows (e.g. an out-of-order spreadsheet row)
        if date.date() > datetime.now().date():
            skipped += 1
            continue
        timestamp = date.strftime("%Y-%m-%dT00:00:00Z")

        opt_mo = str(row[1]).strip() if row[1] is not None else ""
        futures_sym = _infer_cme_soy(opt_mo, date) if opt_mo else ""

        for col_idx, (provider, location) in col_map.items():
            raw = row[col_idx] if col_idx < len(row) else None
            basis = _parse_basis(raw)
            if basis is None:
                continue
            counts[(provider, location)] += 1
            if apply:
                pending.append((timestamp, provider, location, futures_sym, basis))

    print(f"\n{'Location':<45}  {'Weeks':>5}")
    print("-" * 55)
    grand = 0
    for (provider, location), n in sorted(counts.items(), key=lambda x: x[0]):
        print(f"  {provider:<12} {location:<33}  {n:>5}")
        grand += n
    print("-" * 55)
    print(f"  {'TOTAL':<45}  {grand:>5}")
    print()

    if not apply:
        print("Dry run — no data written. Re-run with --apply to import.")
        return

    is_pg = db._use_pg()
    conn = db.get_conn()
    c = conn.cursor()
    BATCH = 500
    try:
        for i, (timestamp, provider, location, futures_sym, basis) in enumerate(pending):
            _upsert_one(conn, c, is_pg, timestamp, provider, location, futures_sym, basis)
            total_snaps += 1
            if total_snaps % BATCH == 0:
                conn.commit()
                print(f"  ... {total_snaps}/{len(pending)} committed", flush=True)
        conn.commit()
    finally:
        conn.close()

    print(f"Done. {total_snaps} snapshot(s) written.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import soybean spot basis history")
    parser.add_argument("--apply", action="store_true",
                        help="Write to the database (default is dry run)")
    parser.add_argument("--pg", action="store_true",
                        help="Use DATABASE_URL / Supabase instead of local SQLite")
    parser.add_argument("--file", default=None,
                        help="Path to Excel file (default: history/Corn basis history .xlsx)")
    args = parser.parse_args()

    if args.pg and not os.getenv("DATABASE_URL"):
        from dotenv import load_dotenv
        load_dotenv()
    if args.pg and not os.getenv("DATABASE_URL"):
        print("ERROR: --pg requires DATABASE_URL to be set.")
        sys.exit(1)

    default_file = Path(__file__).parent / "history" / "Corn basis history .xlsx"
    file_path = Path(args.file) if args.file else default_file

    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}")
        sys.exit(1)

    run(file_path, apply=args.apply, pg=args.pg)
