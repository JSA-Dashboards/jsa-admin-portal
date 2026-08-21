"""
import_corn_history.py — Import weekly corn spot basis history from "Corn basis history .xlsx"

Sheet: SA Master
  Col 1  : Date
  Col 5  : Opt Mo  (CME month code letter: Z/H/K/N/U — year inferred from date)
  Col 7+ : One location per column, basis in cents, "Closed"/blank = skip

Only the 48 columns that match currently-scraped locations are imported.
All other columns are ignored.  Grain = Corn, spot basis only.

Usage:
    python import_corn_history.py               # dry run — shows counts only
    python import_corn_history.py --apply       # write to local SQLite
    python import_corn_history.py --apply --pg  # write to Supabase
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import database as db

# ── Column mapping (0-based index → (provider, db_location)) ─────────────────
# Only columns we actively scrape; every other column is ignored.

COLUMN_MAP: dict[int, tuple[str, str]] = {
    6:  ("ADM",       "Cedar Rapids, IA"),
    7:  ("ADM",       "Columbus, NE (Corn Processing)"),
    8:  ("ADM",       "Decatur, IL (Corn Processing)"),
    9:  ("ADM",       "Marshall, MN (Corn Processing)"),
    10: ("ADM",       "St. Louis, MO (Elevator)"),
    13: ("Cargill",   "Albion"),
    14: ("Cargill",   "Blair"),
    22: ("GPRE",      "Central City"),
    23: ("GPRE",      "Shenandoah"),
    24: ("GPRE",      "Wood River"),
    25: ("GPRE",      "York"),
    33: ("LDC",       "Grand Junction"),
    38: ("POET",      "Chancellor, SD"),
    39: ("POET",      "Ashton, IA"),
    40: ("POET",      "Fairmont, NE"),
    41: ("POET",      "Hudson, SD"),
    42: ("POET",      "Jewell, IA"),
    43: ("POET",      "Mitchell, SD"),
    46: ("POET",      "Fairbank, IA"),
    47: ("POET",      "Menlo, IA"),
    48: ("POET",      "Arthur, IA"),
    49: ("POET",      "Emmetsburg, IA"),
    50: ("POET",      "Shell Rock, IA"),
    51: ("POET",      "Iowa Falls, IA"),
    52: ("Cargill",   "Eddyville"),
    62: ("GPRE",      "Superior"),
    65: ("POET",      "Big Stone City, SD"),
    70: ("GPRE",      "Mount Vernon"),
    71: ("CGB",       "CGB MT VERNON"),
    72: ("ADM",       "Evansville, IN (Ohio St.)"),
    73: ("Primient",  "Lafayette"),
    74: ("Cargill",   "Linden"),
    75: ("POET",      "Portland, IN"),
    76: ("POET",      "North Manchester, IN"),
    77: ("POET",      "Fostoria, OH"),
    78: ("POET",      "Leipsic, OH"),
    80: ("POET",      "Marion, OH"),
    81: ("Cargill",   "Dayton"),
    82: ("Cargill",   "Bloomingburg"),
    84: ("Scoular",   "Waverly, IL - Elevator"),
    85: ("ADM",       "Curran, IL"),
    86: ("Primient",  "Decatur"),
    87: ("GPRE",      "Madison"),
    92: ("POET",      "Laddonia, MO"),
    95: ("POET",      "Macon, MO"),
    99: ("Scoular",   "Scoular Butterfield Cash Bids"),
    105: ("Andersons", "Albion"),
    106: ("POET",     "Caro, MI"),
    108: ("ADM",      "Webberville, MI"),
    109: ("ADM",      "Grand Ledge, MI"),
}

# CME month code → calendar month number
_CME_MONTH_NUM: dict[str, int] = {
    "F": 1, "G": 2, "H": 3, "J": 4, "K": 5, "M": 6,
    "N": 7, "Q": 8, "U": 9, "V": 10, "X": 11, "Z": 12,
}


def _infer_cme_corn(opt_mo: str, date: datetime) -> str:
    """
    Build a full CME corn symbol from the Opt Mo letter and the row date.
    e.g.  "Z" + 2020-09-02  →  "ZCZ20"
          "H" + 2020-12-16  →  "ZCH21"  (March hasn't arrived yet in Dec)
    """
    code = opt_mo.strip().upper()
    month_num = _CME_MONTH_NUM.get(code)
    if not month_num:
        return ""
    year = date.year
    if month_num < date.month:
        year += 1
    return f"ZC{code}{year % 100:02d}"


def _parse_basis(raw) -> int | None:
    """Return integer cents or None.  Skips 'Closed', blank, non-numeric."""
    if raw is None:
        return None
    txt = str(raw).strip()
    if not txt or txt.lower() in ("closed", "n/a", "-", "—", "tbd"):
        return None
    try:
        return int(round(float(txt)))
    except (ValueError, TypeError):
        return None


def _upsert_one(conn, c, is_pg: bool, timestamp: str, provider: str, location: str,
                futures_sym: str, basis: int) -> None:
    """Insert a single spot snapshot using an already-open connection."""
    if is_pg:
        c.execute(
            """INSERT INTO snapshots (timestamp, provider, location, source)
               VALUES (%s, %s, %s, %s)
               ON CONFLICT (timestamp, provider, location) DO NOTHING
               RETURNING id""",
            (timestamp, provider, location, "historical"),
        )
        row = c.fetchone()
        if row:
            snap_id = row["id"]
        else:
            c.execute(
                "SELECT id FROM snapshots WHERE timestamp=%s AND provider=%s AND location=%s",
                (timestamp, provider, location),
            )
            snap_id = c.fetchone()["id"]
        c.execute(
            """INSERT INTO snapshot_rows
               (snapshot_id, row_id, grain, delivery_month, futures_symbol,
                basis_cents, is_spot, spot_grain)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
               ON CONFLICT (snapshot_id, row_id) DO NOTHING""",
            (snap_id, "HIST_CORN_SPOT", "Corn", "Corn", futures_sym, basis, 1, "Corn"),
        )
    else:
        c.execute(
            """INSERT OR IGNORE INTO snapshots (timestamp, provider, location, source)
               VALUES (?, ?, ?, ?)""",
            (timestamp, provider, location, "historical"),
        )
        if c.lastrowid == 0:
            c.execute(
                "SELECT id FROM snapshots WHERE timestamp=? AND provider=? AND location=?",
                (timestamp, provider, location),
            )
            snap_id = c.fetchone()["id"]
        else:
            snap_id = c.lastrowid
        c.execute(
            """INSERT OR IGNORE INTO snapshot_rows
               (snapshot_id, row_id, grain, delivery_month, futures_symbol,
                basis_cents, is_spot, spot_grain)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (snap_id, "HIST_CORN_SPOT", "Corn", "Corn", futures_sym, basis, 1, "Corn"),
        )


def run(file_path: Path, apply: bool = False, pg: bool = False) -> None:
    import openpyxl

    backend = "PostgreSQL" if db._use_pg() else "SQLite"
    print(f"\nCorn History Import  [{backend}]  mode={'APPLY' if apply else 'DRY RUN'}")
    print("=" * 60)
    print(f"File: {file_path.name}")

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    sheet_name = "Corn" if "Corn" in wb.sheetnames else "SA Master"
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Rows 0 and 1 are headers; data starts at row index 2
    data_rows = rows[2:]

    # Counters per location
    counts: dict[tuple, int] = {v: 0 for v in COLUMN_MAP.values()}
    skipped = 0
    total_snaps = 0

    db.init_db()

    # Build list of all (timestamp, provider, location, futures_sym, basis) to insert
    pending: list[tuple] = []

    for row in data_rows:
        raw_date = row[0]
        if raw_date is None:
            continue
        if isinstance(raw_date, datetime):
            date = raw_date
        else:
            try:
                date = datetime.strptime(str(raw_date).strip()[:10], "%Y-%m-%d")
            except ValueError:
                skipped += 1
                continue

        # Skip stray future-dated rows (e.g. an out-of-order spreadsheet row)
        if date.date() > datetime.now().date():
            skipped += 1
            continue
        timestamp = date.strftime("%Y-%m-%dT00:00:00Z")

        opt_mo = str(row[4]).strip() if row[4] is not None else ""
        futures_sym = _infer_cme_corn(opt_mo, date) if opt_mo else ""

        for col_idx, (provider, location) in COLUMN_MAP.items():
            raw = row[col_idx] if col_idx < len(row) else None
            basis = _parse_basis(raw)
            if basis is None:
                continue
            counts[(provider, location)] += 1
            if apply:
                pending.append((timestamp, provider, location, futures_sym, basis))

    # Report counts
    print(f"\n{'Location':<45}  {'Weeks':>5}")
    print("-" * 55)
    grand = 0
    for (provider, location), n in sorted(counts.items(), key=lambda x: x[0]):
        print(f"  {provider:<12} {location:<33}  {n:>5}")
        grand += n
    print("-" * 55)
    print(f"  {'TOTAL':<45}  {grand:>5}")
    print()

    if not apply:
        print("Dry run — no data written. Re-run with --apply to import.")
        return

    # Single connection for all inserts
    is_pg = db._use_pg()
    conn = db.get_conn()
    c = conn.cursor()
    BATCH = 500
    try:
        for i, (timestamp, provider, location, futures_sym, basis) in enumerate(pending):
            _upsert_one(conn, c, is_pg, timestamp, provider, location, futures_sym, basis)
            total_snaps += 1
            if total_snaps % BATCH == 0:
                conn.commit()
                print(f"  ... {total_snaps}/{len(pending)} committed", flush=True)
        conn.commit()
    finally:
        conn.close()

    print(f"Done. {total_snaps} snapshot(s) written.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import corn spot basis history")
    parser.add_argument("--apply", action="store_true",
                        help="Write to the database (default is dry run)")
    parser.add_argument("--pg", action="store_true",
                        help="Use DATABASE_URL / Supabase instead of local SQLite")
    parser.add_argument("--file", default=None,
                        help="Path to Excel file (default: history/Corn basis history .xlsx)")
    args = parser.parse_args()

    if args.pg and not os.getenv("DATABASE_URL"):
        from dotenv import load_dotenv
        load_dotenv()
    if args.pg and not os.getenv("DATABASE_URL"):
        print("ERROR: --pg requires DATABASE_URL to be set.")
        sys.exit(1)

    default_file = Path(__file__).parent / "history" / "Corn basis history .xlsx"
    file_path = Path(args.file) if args.file else default_file

    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}")
        sys.exit(1)

    run(file_path, apply=args.apply, pg=args.pg)
