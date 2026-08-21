"""
Backfill the archive with the pre-2023 weekly history kept in
`History/FOB History.xlsx` — two "master" time-series sheets (Corn, Beans) that
run weekly from 2018-10 through 2023-04, one row per week.

This file is a completely different shape from the per-day-tab workbooks the
other importers read, so it gets its own parser:

  * one row per date; col A = date, col B = front contract letter, col C = CIF
    basis (cents -> $/bu), cols I-M = barge freight per reach (tariff multiplier)
  * Corn and Beans carry their OWN front contract on a given date (e.g. Dec corn
    vs Nov beans), so each commodity's CIF/freight is stored under its own front
    month; freight (commodity-agnostic) is written under both months
  * no CBOT futures / spreads in this file, so these snapshots are CIF+freight
    only — exactly like the archive's other older dates

To avoid clobbering the richer 2023+ snapshots, only dates strictly EARLIER than
the archive's current minimum date are written.

Usage:
  python import_fob_master.py            # DRY RUN — report only, writes nothing
  python import_fob_master.py --commit   # write to the archive (DATABASE_URL)

Follows the project rule: for a real --commit, DATABASE_URL must be set (Postgres)
so the shared archive is updated, never the local SQLite fallback.
"""
import os
import sys
import warnings
import datetime as dt

warnings.filterwarnings("ignore")
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
except ImportError:
    pass

import openpyxl

import db

HISTORY_XLSX = os.environ.get("FOB_HISTORY_XLSX") or (
    r"C:\Users\KoltenPostin\John Stewart and Associates\JSA - Documents"
    r"\St. Louis\JSA FOB Sheet\History\FOB History.xlsx")

SHEETS = {"Corn FOB Master": "Corn", "Bean FOB Master": "Soybeans"}

# Futures contract month letter -> month key (June/July spelled out to match the
# archive's existing convention; the rest 3-letter).
LETTER_MONTH = {"F": "Jan", "G": "Feb", "H": "Mar", "J": "Apr", "K": "May",
                "M": "June", "N": "July", "Q": "Aug", "U": "Sep", "V": "Oct",
                "X": "Nov", "Z": "Dec"}

# Master barge-freight columns (1-indexed) -> canonical archive reach(es).
# Lower Miss has no column here, so Greenville/Memphis/Cairo stay blank pre-2023.
FREIGHT_COLS = {
    9:  ["Upper Miss"],                          # Cities BF
    10: ["Davenport South", "McGregor South"],   # MM BF
    11: ["STL"],                                 # STL BF
    12: ["Ohio"],                                # Lower OHIO BF
    13: ["IL"],                                  # ILL BF
}

CIF_COL, LETTER_COL = 3, 2


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _read_sheet(ws):
    """-> {date: {'month':.., 'cif':.., 'freight':{reach: mult}}} for usable rows."""
    out = {}
    for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        d = row[0]
        if not isinstance(d, dt.datetime):
            continue
        letter = str(row[LETTER_COL - 1] or "").strip().upper()
        month = LETTER_MONTH.get(letter)
        cif = _num(row[CIF_COL - 1])
        if not month or cif is None:
            continue
        freight = {}
        for col, reaches in FREIGHT_COLS.items():
            mult = _num(row[col - 1])
            if mult is None:
                continue
            for reach in reaches:
                freight[reach] = mult
        if not freight:
            continue
        out[d.date()] = {"month": month, "letter": letter,
                         "cif": round(cif / 100.0, 4), "freight": freight}
    return out


def load_master():
    wb = openpyxl.load_workbook(HISTORY_XLSX, read_only=True, data_only=True)
    per_commodity = {}
    for sheet, commodity in SHEETS.items():
        if sheet in wb.sheetnames:
            per_commodity[commodity] = _read_sheet(wb[sheet])
    # merge by date into archive snapshots
    all_dates = sorted(set().union(*[set(v) for v in per_commodity.values()]))
    snapshots = {}
    for d in all_dates:
        cif, calendar, freight = {}, {}, {}
        for commodity, rows in per_commodity.items():
            rec = rows.get(d)
            if not rec:
                continue
            m = rec["month"]
            cif[commodity] = {m: rec["cif"]}
            calendar[commodity] = [(m, rec["letter"])]
            # freight (commodity-agnostic) stored under this commodity's month too
            for reach, mult in rec["freight"].items():
                freight.setdefault(reach, {})[m] = mult
        if cif:
            snapshots[d] = (cif, freight, calendar)
    return snapshots


def run(commit):
    snapshots = load_master()
    if not snapshots:
        print("No usable rows parsed from the master history.")
        return
    lo, hi = min(snapshots), max(snapshots)
    print(f"Parsed {len(snapshots)} weekly snapshots from the master history "
          f"({lo} -> {hi}).")

    existing = db.list_dates()
    existing_dates = {dt.date.fromisoformat(s) for s in existing}
    cutoff = min(existing_dates) if existing_dates else hi + dt.timedelta(days=1)
    print(f"Archive backend: {db.describe_backend() if hasattr(db, 'describe_backend') else ('Postgres' if db._is_postgres() else 'SQLite')}")
    print(f"Archive currently holds {len(existing_dates)} dates"
          + (f" (earliest {min(existing_dates)})." if existing_dates else "."))

    to_write = sorted(d for d in snapshots
                      if d < cutoff and d not in existing_dates)
    skipped = len(snapshots) - len(to_write)
    print(f"Cutoff (archive minimum): {cutoff} — importing dates strictly before it.")
    print(f"Would write {len(to_write)} new snapshots "
          f"({to_write[0]} -> {to_write[-1]} )." if to_write
          else "Nothing new to write.")
    print(f"Skipping {skipped} dates (>= cutoff or already archived).")

    # eyeball a few
    for d in (to_write[:2] + to_write[-2:]) if to_write else []:
        cif, freight, cal = snapshots[d]
        c = {k: v for k, v in cif.items()}
        print(f"  {d}: CIF={c}  freight_reaches={sorted(freight)}  cal={cal}")

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit to archive.")
        return

    if not db._is_postgres():
        print("\nREFUSING to commit: DATABASE_URL is not set, so this would write "
              "to the local SQLite fallback instead of the shared archive. Set "
              "DATABASE_URL (Supabase) and re-run with --commit.")
        sys.exit(1)

    db.init_db()
    written = 0
    for d in to_write:
        cif, freight, calendar = snapshots[d]
        db.save_snapshot(d.isoformat(), cif, freight, calendar)
        written += 1
    now = db.list_dates()
    print(f"\nCommitted {written} snapshots. Archive now spans "
          f"{min(now)} -> {max(now)} ({len(now)} dates).")


if __name__ == "__main__":
    run(commit="--commit" in sys.argv)
